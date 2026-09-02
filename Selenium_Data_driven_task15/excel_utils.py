import os
from datetime import datetime
from openpyxl import load_workbook

EXCEL_FILE = "login_data.xlsx"
SHEET_NAME = "LoginTests"   # change to your actual sheet name if different

class ExcelManager:
    """Reads login data from an existing Excel file and writes results back."""

    def __init__(self, file_path=EXCEL_FILE, sheet_name=SHEET_NAME):
        self.file_path = "C:/Users/balagars/Desktop/training/Selenium/login_data.xlsx"
        self.sheet_name = sheet_name
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(
                f"Excel file '{self.file_path}' not found. "
                f"Place your login data file next to the test code."
            )

    def _get_sheet(self, wb):
        # fall back to the active sheet if the named one isn't present
        if self.sheet_name in wb.sheetnames:
            return wb[self.sheet_name]
        return wb.active

    def read_login_data(self):
        """Return a list of dicts, one per data row (header row skipped)."""
        wb = load_workbook(self.file_path)
        ws = self._get_sheet(wb)
        rows = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:      # stop/skip on empty Test ID
                continue
            rows.append({
                "row": idx,
                "test_id": row[0],
                "username": row[1] if row[1] is not None else "",
                "password": row[2] if row[2] is not None else "",
                "tester": row[5] if len(row) > 5 and row[5] is not None else "",
            })
        wb.close()
        return rows

    def write_result(self, row_index, result):
        """Write Date (col 4), Time (col 5) and Test Result (col 7) back."""
        wb = load_workbook(self.file_path)
        ws = self._get_sheet(wb)
        now = datetime.now()
        ws.cell(row=row_index, column=4, value=now.strftime("%Y-%m-%d"))   # Date
        ws.cell(row=row_index, column=5, value=now.strftime("%H:%M:%S"))   # Time of Test
        ws.cell(row=row_index, column=7, value=result)                     # Test Result
        wb.save(self.file_path)
        wb.close()
